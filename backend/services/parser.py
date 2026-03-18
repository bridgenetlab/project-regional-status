"""Excel file parsing and data normalization."""

from datetime import datetime, timedelta
from typing import Optional
from openpyxl import load_workbook
from io import BytesIO


# Required columns (case-insensitive)
REQUIRED_COLUMNS = {
    'no', 'branch_name', 'status', 'planned_start_date', 'target_end_date',
    'actual_end_date', 'expected_delivery', 'delivery_status', 'site_readiness',
    'installation_date', 'region', 'address', 'pic', 'voice_gateway',
    'cisco_9851', 'cisco_9841', 'poe', 'field_engineer', 'hp', 'ic',
    'voip_ip', 'gateway', 'note', 'physical_phone_installation',
    'issue', 'dependency', 'uat_document', 'installation_cost', 'installation_vendor'
}

# Excel column name aliases (map friendly names to normalized internal keys)
COLUMN_ALIASES = {
    'no': ['no', 'number'],
    'branch_name': ['branch name', 'branch', 'site name'],
    'status': ['status'],
    'planned_start_date': ['planned start date', 'start date', 'planned date'],
    'target_end_date': ['target end date', 'end date', 'target date'],
    'actual_end_date': ['actual end date', 'completion date', 'actual date'],
    'expected_delivery': ['expected delivery', 'delivery date'],
    'delivery_status': ['delivery status', 'delivery'],
    'site_readiness': ['site readiness', 'readiness'],
    'installation_date': ['installation date', 'date installed'],
    'region': ['region', 'location'],
    'address': ['address'],
    'pic': ['pic', 'contact', 'pengurus'],
    'voice_gateway': ['voice gateway', 'gateway'],
    'cisco_9851': ['cisco 9851', 'cisco9851'],
    'cisco_9841': ['cisco 9841', 'cisco9841'],
    'poe': ['poe'],
    'field_engineer': ['field engineer', 'engineer', 'bridgenet pic'],
    'hp': ['hp'],
    'ic': ['ic'],
    'voip_ip': ['voip ip', 'voip_ip'],
    'gateway': ['gateway ip', 'gateway'],
    'note': ['note', 'notes', 'remarks'],
    'physical_phone_installation': ['physical phone installation', 'phone installation'],
    'issue': ['issue', 'issues'],
    'dependency': ['dependency', 'dependencies'],
    'uat_document': ['uat document', 'uat'],
    'installation_cost': ['installation cost', 'cost'],
    'installation_vendor': ['installation vendor', 'vendor']
}


def detect_columns(headers: list[str]) -> dict[str, int]:
    """
    Map Excel headers to internal column keys.
    Returns dict of {internal_key: column_index} for found columns.
    """
    # Normalize headers
    normalized_headers = [h.strip().lower() if h else '' for h in headers]

    column_map = {}
    missing = []

    for internal_key, aliases in COLUMN_ALIASES.items():
        found = False
        for idx, header in enumerate(normalized_headers):
            if header in aliases:
                column_map[internal_key] = idx
                found = True
                break
        # Don't mark as missing - just skip if not found
        # if not found:
        #     missing.append(internal_key)

    return column_map, missing


def parse_date(value: Optional[str | int | float | datetime]) -> Optional[str]:
    """
    Parse various date formats from Excel cell values.
    Returns ISO format string (YYYY-MM-DD) or None if unparseable.
    """
    if value is None or value == '':
        return None

    # Already a datetime object (openpyxl parsed it)
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')

    # Excel float serial (days since 1899-12-30)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            excel_epoch = datetime(1899, 12, 30)
            parsed_date = excel_epoch + timedelta(days=value)
            return parsed_date.strftime('%Y-%m-%d')
        except (ValueError, OverflowError):
            return None

    # String - try common formats
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None

        formats = ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d %b %Y', '%d %b %y']
        for fmt in formats:
            try:
                parsed = datetime.strptime(value, fmt)
                return parsed.strftime('%Y-%m-%d')
            except ValueError:
                continue

        # If all parsing fails, return as-is (marked as unparseable upstream)
        return None

    return None


def parse_excel_file(file_bytes: bytes) -> tuple[list[dict], list[str], list[str]]:
    """
    Parse Excel file.
    Returns (data_rows, columns_detected, warnings).
    """
    warnings = []

    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"Failed to parse Excel file: {str(e)}")

    # Get first sheet
    if not wb.sheetnames:
        raise ValueError("Excel file has no sheets")
    ws = wb[wb.sheetnames[0]]

    # Extract headers from first row
    headers = [cell.value for cell in ws[1]]
    if not any(headers):
        raise ValueError("Excel file has no headers in first row")

    # Detect column mapping
    column_map, missing = detect_columns(headers)
    columns_detected = [h for h in headers if h]

    # Parse data rows
    data_rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):  # Skip empty rows
            continue

        # Initialize record with all fields as None
        record = {k: None for k in REQUIRED_COLUMNS}

        # Extract mapped columns
        for internal_key, col_idx in column_map.items():
            raw_value = row[col_idx] if col_idx < len(row) else None

            # Parse date fields
            if 'date' in internal_key or internal_key == 'installation_date':
                record[internal_key] = parse_date(raw_value)
                if raw_value is not None and record[internal_key] is None:
                    warnings.append(f"Row {row_idx}: Could not parse date in {internal_key}: {raw_value}")
            # Parse numeric fields
            elif internal_key in ['cisco_9851', 'cisco_9841', 'no']:
                try:
                    record[internal_key] = int(raw_value) if raw_value is not None else None
                except (ValueError, TypeError):
                    record[internal_key] = None
            # String fields
            else:
                record[internal_key] = str(raw_value).strip() if raw_value is not None else None

        # Initialize computed fields
        record['days_delta'] = None
        record['schedule_status'] = None

        data_rows.append(record)

    return data_rows, columns_detected, warnings
